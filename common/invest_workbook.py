# -*- coidng:utf-8 -*-
'''
  Author : shuo
  Date : 2019/6/3 11:01
  File : invest_workbook.py
  Project : ClassProjects
 '''
from openpyxl import load_workbook
from API_1.path.file_path import api_path
from API_1.common.connect_db import db_connect
from API_1.common.getData import getData

class load_invest:
    def __init__(self,sheet):
        self.sheet = sheet

    def load_investCase(self):
        workbook = load_workbook(api_path)
        sheet = workbook[self.sheet]
        investcase = []
        for i in range(2,sheet.max_row+1):
            case = {}
            case['Caseid'] = sheet.cell(i,1).value
            case['Moudle'] = sheet.cell(i, 2).value
            case['Title'] = sheet.cell(i, 3).value
            case['Url'] = sheet.cell(i, 4).value
            case['Method'] = sheet.cell(i, 5).value
            case['Params'] = sheet.cell(i, 6).value
            case['sql'] = sheet.cell(i, 7).value
            # if sheet.cell(i,7).value != None:
            #     leaveAmount = db_connect().connect_mysql(eval(sheet.cell(i,7).value)['sql'])
            #     setattr(getData,'leaveAmount',leaveAmount)
            case['ExpectedResult'] = sheet.cell(i, 8).value
            investcase.append(case)
        return investcase
    def write_result(self,row,column,value):
        workbook = load_workbook(api_path)
        sheet = workbook[self.sheet]
        sheet.cell(row,column,value)
        workbook.save(api_path)
        workbook.close()

if __name__ == '__main__':
    print(load_invest('invest').load_investCase())
