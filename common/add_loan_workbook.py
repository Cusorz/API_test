# -*- coidng:utf-8 -*-
'''
  Author : shuo
  Date : 2019/5/31 9:44
  File : add_loan_workbook.py
  Project : ClassProjects
 '''
from openpyxl import load_workbook
from API_1.path.file_path import api_path
from API_1.common.connect_db import db_connect
from API_1.common.getData import getData

class add_loan:
    def __init__(self,sheet):
        self.sheet = sheet
    def load_loanCase(self):
        workbook = load_workbook(api_path)
        sheet = workbook[self.sheet]
        loanCase = []
        for i in range(2,sheet.max_row+1):
            case = {}
            case['Caseid'] = sheet.cell(i,1).value
            case['Moudle'] = sheet.cell(i, 2).value
            case['Title'] = sheet.cell(i, 3).value
            case['Url'] = sheet.cell(i, 4).value
            case['Method'] = sheet.cell(i, 5).value
            case['Params'] = sheet.cell(i, 6).value
            #如果sql一列不为空，则执行sql语句，并把loanid保存，需要时通过反射获取值
            if sheet.cell(i,7).value != None:
                loanid = db_connect().connect_mysql(eval(sheet.cell(i,7).value)['sql'])
                setattr(getData,'loanid',loanid[0])
            case['ExpectedResult'] = sheet.cell(i, 8).value
            loanCase.append(case)
        return loanCase

    def write_result(self,row,column,value):
        workbook = load_workbook(api_path)
        sheet = workbook[self.sheet]
        sheet.cell(row,column,value)
        workbook.save(api_path)
        workbook.close()

if __name__ == '__main__':
    print(add_loan('addloan').load_loanCase())