# -*- coidng:utf-8 -*-
'''
  Author : shuo
  Date : 2019/5/30 9:20
  File : workbook.py
  Project : ClassProjects
 '''
from openpyxl import load_workbook
from API_1.path.file_path import api_path
from API_1.common.read_conf import readConf

conf = readConf().read_caseconf('registerConf','caseConf')
# print(conf)
class workBook:
    def __init__(self,sheet):
        self.sheet = sheet
    #加载测试用例并存在列表里面
    def loading_case(self):
        workbook = load_workbook(api_path)
        sheet = workbook[self.sheet]
        allcase = []
        for i in range(2,sheet.max_row+1):
            case = {}
            case['Caseid'] = sheet.cell(i,1).value
            case['Moudle'] = sheet.cell(i, 2).value
            case['Title'] = sheet.cell(i, 3).value
            case['Url'] = sheet.cell(i, 4).value
            case['Method'] = sheet.cell(i, 5).value
            # case['Params'] = sheet.cell(i, 6).value
            if sheet.cell(i,6).value.find('tel') != -1:
                case['Params'] = sheet.cell(i,6).value.replace('tel',str(self.getphone()))
                self.update_phone()
            else:
                case['Params'] = sheet.cell(i,6).value
            case['sql'] = sheet.cell(i,7).value
            case['ExpectedResult'] = sheet.cell(i,8).value
            allcase.append(case)
        final_data = []
        if conf == 'all':
            final_data = allcase
        else:
            for x in conf:
                #这里final_data也是个列表，可以直接通过数组下标添加allcase下符合条件的用例,下面这个是错误写法：final_data = allcase.append(allcase[x-1])
                final_data.append(allcase[x-1])
        workbook.close()
        return final_data
    #把测试结果回填到excel表格中
    def write_result(self,row,cloumn,value):
        workBook = load_workbook(api_path)
        sheet = workBook[self.sheet]
        sheet.cell(row,cloumn,value)
        workBook.save(api_path)
        workBook.close()
    #获取变量值，并把获取到的值返回，提供给其他调用者使用
    def getphone(self):
        workBook = load_workbook(api_path)
        sheet = workBook['phone']
        phone = sheet.cell(1,2).value
        return phone
    #每次调用完都把手机号进行更新操作
    def update_phone(self):
        workBook = load_workbook(api_path)
        sheet = workBook['phone']
        phone = self.getphone()
        #这里不需要替换操作，可直接更新手机号
        sheet.cell(1,2,int(phone)+1)
        workBook.save(api_path)
        workBook.close()

if __name__ == '__main__':
    print(workBook('recharge').loading_case())
